import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook('Шаблон.xlsx')
ws = wb.active

# Check rows 21-26 - delivery section and grand totals
print('=== ROWS 20-26 ===')
for r in range(20, 27):
    val = str(ws.cell(row=r, column=1).value or '').strip()
    h = ws.row_dimensions[r].height
    c1 = ws.cell(row=r, column=1)
    c5 = ws.cell(row=r, column=5)
    fill1 = c1.fill.fgColor.rgb if c1.fill else 'none'
    font1 = f'{c1.font.name},{c1.font.size},bold={c1.font.bold}' if c1.font else 'none'
    font5 = f'{c5.font.name},{c5.font.size},bold={c5.font.bold}' if c5.font else 'none'
    print(f'Row {r} (h={h}): "{val[:40]}" | fill1={fill1} | font1={font1} | font5={font5}')

# Also check data row height and wrap_text
print()
print('=== DATA ROW 7 ===')
c = ws.cell(row=7, column=1)
print(f'fill={c.fill.fgColor.rgb}')
print(f'wrap_text={c.alignment.wrap_text}')
print(f'height={ws.row_dimensions[7].height}')
