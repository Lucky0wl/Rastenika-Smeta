from openpyxl import load_workbook
import os

def analyze_diff():
    gen_path = r'c:\Users\Server\Desktop\Rastenika Smeta\Смета_Ландшафт_1774956019644.xlsx'
    
    wb_gen = load_workbook(gen_path)
    ws_gen = wb_gen.active
    wb_tmpl = load_workbook('Шаблон.xlsx')
    ws_tmpl = wb_tmpl.active

    print("--- Generated File Footer Analysis ---")
    for r in range(1, ws_gen.max_row + 1):
        val = ws_gen.cell(r, 1).value
        val5 = ws_gen.cell(r, 5).value 
        if val and isinstance(val, str) and 'оплате' in val.lower():
            print(f"Row {r}: '{val}'")
            print(f"Row {r} heights: Gen={ws_gen.row_dimensions[r].height}")
        if val5 and isinstance(val5, str) and 'оплате' in val5.lower():
            print(f"Found 'оплате' in Col 5, Row {r}")

    print("\n--- Generated File Item Styling ---")
    c = ws_gen.cell(7, 1)
    if c.font:
        print(f"Row 7 Item 1: Font={c.font.name} {c.font.size}, Alignment={c.alignment.horizontal}/{c.alignment.vertical}, Wrap={c.alignment.wrap_text}")
    
    print("\n--- Template Row 7 Styling ---")
    c_t = ws_tmpl.cell(7, 1)
    if c_t.font:
        print(f"Row 7 Item 1: Font={c_t.font.name} {c_t.font.size}, Alignment={c_t.alignment.horizontal}/{c_t.alignment.vertical}, Wrap={c_t.alignment.wrap_text}")

if __name__ == '__main__':
    analyze_diff()
