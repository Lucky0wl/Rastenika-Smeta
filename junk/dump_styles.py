"""
Debug: dump full cell styles of Шаблон.xlsx rows 1-30 to understand all styles
we need to replicate.
"""
from openpyxl import load_workbook

wb = load_workbook('Шаблон.xlsx')
ws = wb.active

def cell_full(cell):
    try:
        fill_rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb' else 'none'
        font_info = f"name={cell.font.name},size={cell.font.size},bold={cell.font.bold},color={cell.font.color.rgb if cell.font.color and cell.font.color.type == 'rgb' else 'none'}"
        return f"fill={fill_rgb} | font=({font_info}) | h_align={cell.alignment.horizontal if cell.alignment else 'none'}"
    except:
        return 'err'

print("=== TEMPLATE STYLES ===")
for r in [5, 6, 7, 8, 53, 55, 56, 57, 58, 59, 60, 61, 65, 66, 67, 68, 69, 70]:
    for c in range(1, 6):
        cell = ws.cell(row=r, column=c)
        print(f"R{r}C{c}: val={repr(cell.value)} | {cell_full(cell)}")
    print()

print("Row heights:", {r: ws.row_dimensions[r].height for r in [5,6,7,8,53,55,56,57,58,59,60,65,66,67,68,69,70]})
print("Column widths:", {c: ws.column_dimensions[c].width for c in ['A','B','C','D','E']})
