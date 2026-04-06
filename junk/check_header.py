import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook('Шаблон.xlsx')
ws = wb.active
print('=== TEMPLATE rows 1-6 ===')
for r in range(1, 7):
    for col in range(1, 6):
        c = ws.cell(row=r, column=col)
        align = f'H:{c.alignment.horizontal} V:{c.alignment.vertical}' if c.alignment else '-'
        fill = c.fill.fgColor.rgb if c.fill else 'none'
        merged_info = ''
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= r <= mr.max_row and mr.min_col <= col <= mr.max_col:
                merged_info = f' [merge {mr.coord}]'
        print(f'  R{r}C{col}: val={repr(str(c.value)[:30])} align={align} fill={fill} class={type(c).__name__}{merged_info}')
