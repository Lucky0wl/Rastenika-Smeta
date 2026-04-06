"""
Deep compare Шаблон.xlsx vs generated Смета to find ALL layout differences:
- Colors, fonts, borders, merges, row heights, column widths, cell values
"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import json, copy

def cell_info(cell):
    fill = None
    if cell.fill:
        fg = cell.fill.fgColor
        if fg and fg.type == 'rgb':
            fill = fg.rgb
        elif fg and fg.type == 'indexed':
            fill = f"indexed:{fg.indexed}"
    return {
        'value': cell.value,
        'fill': fill,
        'font_bold': cell.font.bold if cell.font else None,
        'font_color': cell.font.color.rgb if (cell.font and cell.font.color and cell.font.color.type == 'rgb') else None,
        'font_size': cell.font.size if cell.font else None,
        'font_name': cell.font.name if cell.font else None,
        'alignment_h': cell.alignment.horizontal if cell.alignment else None,
        'alignment_wrap': cell.alignment.wrap_text if cell.alignment else None,
        'number_format': cell.number_format,
    }

def dump_sheet(ws, name, max_row=80):
    rows = []
    for r in range(1, max_row + 1):
        row_info = {
            'row': r,
            'height': ws.row_dimensions[r].height,
            'cells': {}
        }
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            row_info['cells'][c] = cell_info(cell)
        rows.append(row_info)
    return rows

wb_t = load_workbook('Шаблон.xlsx')
ws_t = wb_t.active

latest = sorted([f for f in __import__('os').listdir('temp_pdfs') if f.endswith('.xlsx')],
                key=lambda f: __import__('os').path.getmtime('temp_pdfs/' + f))
if not latest:
    # try generated files
    latest = sorted([f for f in __import__('os').listdir('.') if f.startswith('Смета') and f.endswith('.xlsx')],
                    key=lambda f: __import__('os').path.getmtime(f))

if latest:
    gen_file = 'temp_pdfs/' + latest[-1] if __import__('os').path.exists('temp_pdfs/' + latest[-1]) else latest[-1]
else:
    gen_file = 'Смета_Ландшафт_1775118715805.xlsx'

print(f"Comparing template vs: {gen_file}")
wb_g = load_workbook(gen_file)
ws_g = wb_g.active

tmpl = dump_sheet(ws_t, "Template")
genr = dump_sheet(ws_g, "Generated")

report = []
for i, (tr, gr) in enumerate(zip(tmpl[:30], genr[:30])):
    if tr['height'] != gr['height']:
        report.append(f"Row {tr['row']}: height template={tr['height']} vs gen={gr['height']}")
    for c in range(1, 7):
        tc = tr['cells'][c]
        gc = gr['cells'][c]
        diffs = []
        for k in ['fill', 'font_bold', 'font_color', 'font_size', 'font_name', 'alignment_h']:
            if tc[k] != gc[k]:
                diffs.append(f"{k}: T={tc[k]} G={gc[k]}")
        if diffs:
            report.append(f"Row {tr['row']} Col {c}: {'; '.join(diffs)}")

# Column widths
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
    tw = ws_t.column_dimensions[col_letter].width
    gw = ws_g.column_dimensions[col_letter].width
    if abs(tw - gw) > 0.1:
        report.append(f"Column {col_letter}: template width={tw} vs generated={gw}")

# Merges in template
print("Template merges:", [str(m) for m in ws_t.merged_cells.ranges])
print("Generated merges:", [str(m) for m in ws_g.merged_cells.ranges])

with open('compare_report.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(report))

print("Done. Differences:")
print('\n'.join(report[:50]))
