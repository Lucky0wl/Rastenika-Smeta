import os
from openpyxl import load_workbook
import app

data = {
    'items': [{'name': 'Test Plant', 'parameters': 'c2', 'quantity': 1, 'price': 1000, 'planting': 0, 'total': 1000}],
    'material_total': 0,
    'tax_rate': 0
}

try:
    # 1. Check template column widths
    wb_temp = load_workbook('Шаблон.xlsx')
    ws_temp = wb_temp.active
    temp_widths = {}
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        temp_widths[col_letter] = ws_temp.column_dimensions[col_letter].width
    
    print("Template Column Widths:", temp_widths)
    
    # 2. Generate file
    app.app.config['TEMP_FOLDER'] = 'temp_pdfs'
    filepath = app.create_app_xlsx(data)
    
    wb_gen = load_workbook(filepath)
    ws_gen = wb_gen.active
    
    # 3. Check generated column widths
    gen_widths = {}
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        gen_widths[col_letter] = ws_gen.column_dimensions[col_letter].width
        
    print("Generated Column Widths:", gen_widths)
    
    match = True
    for col, width in temp_widths.items():
        if abs(width - gen_widths.get(col, 0)) > 0.1:
            print(f"Width mismatch in column {col}: {width} vs {gen_widths.get(col)}")
            match = False
            
    if match:
        print("Column Widths SUCCESS")
    else:
        print("Column Widths FAILURE")
        
    # 4. Check extended styling (column 10)
    # Plant row is 7
    cell_j7 = ws_gen.cell(row=7, column=10)
    fill_j7 = cell_j7.fill.fgColor.rgb if cell_j7.fill and cell_j7.fill.fgColor else "None"
    print(f"Cell J7 (col 10) fill: {fill_j7}")
    if fill_j7 == 'FF434343':
        print("Extended styling SUCCESS")
    else:
        print("Extended styling FAILURE")

except Exception as e:
    print(f"Error: {e}")
