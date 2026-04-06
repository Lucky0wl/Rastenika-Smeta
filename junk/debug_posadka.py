import sys, io, os, glob
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

print('=== TEMPLATE rows 11-17 (Посадка section) C1-5 ===')
wb = load_workbook('Шаблон.xlsx')
ws = wb.active
for r in range(11, 18):
    for col in range(1, 6):
        c = ws.cell(row=r, column=col)
        fill = c.fill.fgColor.rgb if c.fill else 'none'
        fill_type = c.fill.fill_type if c.fill else '-'
        merged_info = ''
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= r <= mr.max_row and mr.min_col <= col <= mr.max_col:
                merged_info = f' [merge {mr}]'
        print(f'  T R{r}C{col}: val={repr(str(c.value)[:20])} fill={fill} type={fill_type} class={type(c).__name__}{merged_info}')

files = sorted(glob.glob(r'temp_pdfs\*.xlsx'), key=os.path.getmtime, reverse=True)
wb2 = load_workbook(files[0])
ws2 = wb2.active
print()
print('=== GENERATED rows 11-17 (Посадка section) ===')
for r in range(11, 18):
    for col in range(1, 6):
        c = ws2.cell(row=r, column=col)
        fill = c.fill.fgColor.rgb if c.fill else 'none'
        fill_type = c.fill.fill_type if c.fill else '-'
        merged_info = ''
        for mr in ws2.merged_cells.ranges:
            if mr.min_row <= r <= mr.max_row and mr.min_col <= col <= mr.max_col:
                merged_info = f' [merge {mr}]'
        print(f'  G R{r}C{col}: val={repr(str(c.value)[:20])} fill={fill} type={fill_type} class={type(c).__name__}{merged_info}')
