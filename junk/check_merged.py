import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

wb = load_workbook('Шаблон.xlsx')
ws = wb.active

print('=== MERGED CELLS (rows 8-30) ===')
for mr in ws.merged_cells.ranges:
    if mr.min_row >= 8 and mr.min_row <= 30:
        val = ws.cell(row=mr.min_row, column=mr.min_col).value
        print(f'  Rows {mr.min_row}-{mr.max_row}, Cols {mr.min_col}-{mr.max_col}: {repr(str(val or "")[:40])}')

print()
print('=== ROW 10 (plant total) - all cols ===')
for col in range(1, 8):
    c = ws.cell(row=10, column=col)
    fill = c.fill.fgColor.rgb if c.fill and c.fill.fgColor else 'none'
    ctype = type(c).__name__
    val = c.value
    print(f'  Col {col}: fill={fill} type={ctype} val={repr(val)}')
