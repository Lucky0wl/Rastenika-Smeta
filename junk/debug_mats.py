import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# Check template rows 17-20 (materials section data rows)
print('=== TEMPLATE (not data_only) rows 16-21 ===')
wb = load_workbook('Шаблон.xlsx')
ws = wb.active
for r in range(16, 21):
    for col in range(1, 6):
        c = ws.cell(row=r, column=col)
        fill = c.fill.fgColor.rgb if c.fill else 'none'
        fill_type = c.fill.fill_type if c.fill else '-'
        print(f'  Row {r} Col {col}: val={repr(c.value)} fill={fill} filltype={fill_type} type={type(c).__name__}')

print()
# Check latest generated file rows 19-22
import glob, os
files = sorted(glob.glob(r'temp_pdfs\*.xlsx'), key=os.path.getmtime, reverse=True)
wb2 = load_workbook(files[0])
ws2 = wb2.active
print('=== GENERATED rows 19-22 ===')
for r in range(19, 23):
    for col in range(1, 6):
        c = ws2.cell(row=r, column=col)
        fill = c.fill.fgColor.rgb if c.fill else 'none'
        fill_type = c.fill.fill_type if c.fill else '-'
        print(f'  Row {r} Col {col}: val={repr(c.value)} fill={fill} filltype={fill_type} type={type(c).__name__}')
