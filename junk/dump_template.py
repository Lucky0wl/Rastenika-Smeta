from openpyxl import load_workbook

try:
    wb = load_workbook('Шаблон.xlsx')
    ws = wb.active
    
    print(f"Sheet name: {ws.title}")
    
    # Dump first 100 rows to see labels
    data = []
    for r in range(1, 101):
        row_vals = []
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            fill = cell.fill.start_color.index if cell.fill else "None"
            row_vals.append(f"{val} ({fill})")
        data.append(f"Row {r}: {' | '.join(row_vals)}")
        
    with open('template_dump.txt', 'w', encoding='utf-8') as f:
        f.write('\n'.join(data))
    print("Done. Results in template_dump.txt")
except Exception as e:
    print(f"Error: {e}")
