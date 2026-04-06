import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

wb_ref = load_workbook('Шаблон.xlsx', data_only=True)
ws_ref = wb_ref.active

print('=== wb_ref merged cells ===')
for mr in ws_ref.merged_cells.ranges:
    if mr.min_row >= 8 and mr.min_row <= 27:
        val = ws_ref.cell(row=mr.min_row, column=mr.min_col).value
        c1 = ws_ref.cell(row=mr.min_row, column=mr.min_col)
        fill = c1.fill.fgColor.rgb if c1.fill else 'none'
        print(f'  Rows {mr.min_row}-{mr.max_row}, Cols {mr.min_col}-{mr.max_col}: fill={fill} val={repr(str(val or "")[:30])}')

print()
print('=== wb_ref row 10 all cols ===')
for col in range(1, 6):
    c = ws_ref.cell(row=10, column=col)
    fill = c.fill.fgColor.rgb if c.fill else 'none'
    print(f'  Col {col}: fill={fill} type={type(c).__name__}')

print()
print('=== wb (not data_only) row 10 all cols ===')
wb = load_workbook('Шаблон.xlsx')
ws = wb.active
for col in range(1, 6):
    c = ws.cell(row=10, column=col)
    fill = c.fill.fgColor.rgb if c.fill else 'none'
    print(f'  Col {col}: fill={fill} type={type(c).__name__}')
