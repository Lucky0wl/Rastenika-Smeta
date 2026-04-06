import sys, io, os, glob
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

files = sorted(glob.glob(r'temp_pdfs\*.xlsx'), key=os.path.getmtime, reverse=True)
latest = files[0]
print(f'Checking: {latest}')

wb = load_workbook(latest)
ws = wb.active

print()
print('=== ALL ROWS WITH TEXT (rows 7-32) ===')
for r in range(7, 33):
    c1 = ws.cell(row=r, column=1)
    val = str(c1.value or '').strip()[:40]
    fill1 = c1.fill.fgColor.rgb if c1.fill else 'none'
    c2 = ws.cell(row=r, column=2)
    fill2 = c2.fill.fgColor.rgb if c2.fill else 'none'
    c5 = ws.cell(row=r, column=5)
    val5 = str(c5.value or '')
    if val or val5:
        print(f'Row {r}: A="{val}" fill1={fill1} B-fill={fill2} E="{val5[:20]}"')
