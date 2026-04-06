import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import os, glob

# Load latest generated estimate
files = glob.glob(r'temp_pdfs\*.xlsx')
latest = max(files, key=os.path.getmtime)
print(f'Checking: {latest}')

wb = load_workbook(latest)
ws = wb.active

print()
print('=== ИТОГО rows in generated file ===')
for r in range(1, 60):
    val = str(ws.cell(row=r, column=1).value or '').lower()
    if 'итого' in val or 'итог' in val:
        print(f'\nRow {r}: "{ws.cell(row=r, column=1).value}"')
        for col in range(1, 6):
            c = ws.cell(row=r, column=col)
            fill = c.fill.fgColor.rgb if c.fill else 'none'
            print(f'  Col {col}: fill={fill} type={type(c).__name__}')

print()
print('=== MERGED CELLS in generated file ===')
for mr in ws.merged_cells.ranges:
    if mr.min_row >= 7 and mr.min_row <= 40:
        c = ws.cell(row=mr.min_row, column=mr.min_col)
        fill = c.fill.fgColor.rgb if c.fill else 'none'
        print(f'  Rows {mr.min_row}-{mr.max_row}, Cols {mr.min_col}-{mr.max_col}: fill={fill}')
