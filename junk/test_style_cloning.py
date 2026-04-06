"""
Final test: generate XLSX and verify it matches template styles.
"""
import os
os.makedirs('temp_pdfs', exist_ok=True)

import app
from openpyxl import load_workbook

app.app.config['TEMP_FOLDER'] = 'temp_pdfs'

data = {
    'items': [
        {'name': 'Туя западная Aureospicata', 'parameters': '180-200 rb', 'quantity': 1, 'price': 16875, 'planting': 0, 'total': 16875},
        {'name': 'Туя восточная Justynka', 'parameters': '100-120 C25', 'quantity': 1, 'price': 13500, 'planting': 0, 'total': 13500},
    ],
    'material_total': 0,
    'tax_rate': 0
}

try:
    filepath = app.create_app_xlsx(data)
    print(f"Generated: {filepath}")
    
    wb_gen = load_workbook(filepath)
    ws_gen = wb_gen.active
    
    wb_tmpl = load_workbook('Шаблон.xlsx')
    ws_tmpl = wb_tmpl.active
    
    def cell_style_summary(ws, row, col):
        cell = ws.cell(row=row, column=col)
        try:
            fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb' else 'none'
        except:
            fill = 'err'
        font_name = cell.font.name if cell.font else 'none'
        font_size = cell.font.size if cell.font else 'none'
        bold = cell.font.bold if cell.font else 'none'
        try:
            color = cell.font.color.rgb if cell.font and cell.font.color and cell.font.color.type == 'rgb' else 'none'
        except:
            color = 'err'
        return f"fill={fill} | font={font_name},{font_size},bold={bold},color={color}"
    
    print("\n=== Row 7 (plant row) comparison ===")
    for r_template, r_gen in [(7, 7), (7, 8)]:
        print(f"Template R{r_template}C1: {cell_style_summary(ws_tmpl, r_template, 1)}")
        print(f"Generated R{r_gen}C1:  {cell_style_summary(ws_gen, r_gen, 1)}")
        mismatch = cell_style_summary(ws_tmpl, r_template, 1) != cell_style_summary(ws_gen, r_gen, 1)
        print(f"  => {'MISMATCH' if mismatch else 'MATCH'}")
        print()
    
    # Check row heights
    print(f"Template R7 height: {ws_tmpl.row_dimensions[7].height}")
    print(f"Generated R7 height: {ws_gen.row_dimensions[7].height}")
    print(f"Generated R8 height: {ws_gen.row_dimensions[8].height}")
    
except Exception as e:
    import traceback
    traceback.print_exc()
