import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook('Шаблон.xlsx')
ws = wb.active

print('=== All non-empty rows 7-30 ===')
for r in range(7, 31):
    val = str(ws.cell(row=r, column=1).value or '').strip()
    if val:
        print(f'Row {r}: "{val[:50]}"')

# Simulate key_rows scan
print()
print('=== key_rows scan simulation ===')
DATA_START_ROW = 7
key_rows = {}
scan_state = 'looking_plant_total'

for r in range(DATA_START_ROW, 300):
    val = str(ws.cell(row=r, column=1).value or '').lower().strip()
    if not val:
        continue
    
    if scan_state == 'looking_plant_total':
        if 'итого' in val:
            key_rows['plant_total'] = r
            scan_state = 'looking_planting_section'
            print(f'plant_total = {r}: "{val[:40]}"')
    elif scan_state == 'looking_planting_section':
        if 'посадк' in val or 'работ' in val:
            scan_state = 'looking_planting_total'
    elif scan_state == 'looking_planting_total':
        if 'итого' in val:
            key_rows['planting_total'] = r
            scan_state = 'looking_materials_section'
            print(f'planting_total = {r}: "{val[:40]}"')
    elif scan_state == 'looking_materials_section':
        if 'материал' in val:
            scan_state = 'looking_materials_total'
    elif scan_state == 'looking_materials_total':
        if 'итого' in val:
            key_rows['materials_total'] = r
            scan_state = 'looking_grand_total'
            print(f'materials_total = {r}: "{val[:40]}"')
    elif scan_state == 'looking_grand_total':
        if 'итого' in val:
            key_rows['grand_total'] = r
            print(f'grand_total = {r}: "{val[:40]}"')

print()
print('key_rows final:', key_rows)
print()
print('Rows NOT in key_rows that have ИТОГО:')
for r in range(7, 31):
    val = str(ws.cell(row=r, column=1).value or '').lower().strip()
    if 'итого' in val and r not in key_rows.values():
        print(f'  Row {r}: "{val[:50]}"')
