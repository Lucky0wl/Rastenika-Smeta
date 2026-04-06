import os
import uuid
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pdf_generator import PDFGenerator

def test_generate_xlsx(data):
    BASE_DIR = os.path.abspath(os.path.dirname(__file__))
    template_path = os.path.join(BASE_DIR, 'Шаблон.xlsx')
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    items = data.get('items', [])
    if len(items) > 1:
        ws.insert_rows(7, amount=len(items)-1)
        
    for i, item in enumerate(items):
        row_idx = 7 + i
        ws.cell(row=row_idx, column=1).value = item.get('name')
        ws.cell(row=row_idx, column=2).value = item.get('parameters', '')
        ws.cell(row=row_idx, column=3).value = item.get('quantity')
        ws.cell(row=row_idx, column=4).value = item.get('price', 0) + item.get('planting', 0)
        ws.cell(row=row_idx, column=5).value = item.get('total')
        
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = Font(name='Montserrat', size=11)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col == 1: cell.alignment = Alignment(horizontal='left', vertical='center')
            if col >= 4: cell.number_format = '#,##0'

    def write_to_cell(ws, row, col, val):
        coord = f"{get_column_letter(col)}{row}"
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = val
                return
        ws.cell(row=row, column=col).value = val

    offset = max(0, len(items) - 1)
    for r in range(50 + offset, 80 + offset):
        label = str(ws.cell(r, 1).value or "")
        if 'материал' in label.lower():
            write_to_cell(ws, r, 5, data.get('material_total'))
        elif 'работ' in label.lower() or 'посадк' in label.lower():
            write_to_cell(ws, r, 5, data.get('labor_total'))
        elif 'налог' in label.lower() or 'ндс' in label.lower():
            write_to_cell(ws, r, 5, data.get('tax_amount'))
        elif 'итого к оплате' in label.lower() or 'всего' in label.lower():
            write_to_cell(ws, r, 5, data.get('grand_total'))

    out_path = 'final_fidelity_test.xlsx'
    wb.save(out_path)
    return out_path

data = {
    'items': [
        {'name': 'Test Item 1', 'parameters': 'High Fidelity', 'quantity': 10, 'price': 1000, 'planting': 200, 'total': 12000},
        {'name': 'Test Item 2', 'parameters': '1-to-1 Match', 'quantity': 5, 'price': 2000, 'planting': 300, 'total': 11500}
    ],
    'material_total': 20000,
    'labor_total': 3500,
    'tax_rate': 0,
    'tax_amount': 0,
    'grand_total': 23500
}

# Тест XLSX
xlsx_path = test_generate_xlsx(data)
print(f"XLSX Generated successfully at {xlsx_path}")

# Тест PDF
pdf_gen = PDFGenerator()
pdf_path = pdf_gen.create_estimate(data, os.path.abspath('final_fidelity_test.pdf'))
print(f"PDF Generated successfully at {pdf_path}")
