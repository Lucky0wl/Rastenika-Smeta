import os
import uuid
import pandas as pd
from openpyxl import load_workbook
import app # Import the app to use its functions

# Mocking Flask app context if needed, but create_app_xlsx only needs data
# and BASE_DIR. In app.py BASE_DIR is absolute.

data = {
    'items': [
        {'name': 'Туя западная Aureospicata', 'parameters': '180-200 rb', 'quantity': 1, 'price': 16875, 'planting': 0, 'total': 16875},
        {'name': 'Туя восточная Justynka', 'parameters': '100-120 C25', 'quantity': 1, 'price': 13500, 'planting': 0, 'total': 13500},
        {'name': 'Туя западная Brabant', 'parameters': '180-200', 'quantity': 1, 'price': 13365, 'planting': 0, 'total': 13365}
    ],
    'material_total': 32400,
    'tax_rate': 0
}

try:
    print("Generating test XLSX...")
    # We need to set app.config['TEMP_FOLDER'] for the script
    app.app.config['TEMP_FOLDER'] = 'temp_pdfs'
    os.makedirs('temp_pdfs', exist_ok=True)
    
    filepath = app.create_app_xlsx(data)
    print(f"File generated: {filepath}")
    
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Check merges
    print("\nMerges in generated file:")
    found_footer_merge = False
    for m_range in ws.merged_cells.ranges:
        print(f"Merge: {m_range}")
        # Footer for plants should be around row 10 (7 + 3 items)
        if m_range.min_row == 10 and m_range.min_col == 1 and m_range.max_col == 4:
            found_footer_merge = True
            
    if found_footer_merge:
        print("Found restored footer merge A10:D10! SUCCESS")
    else:
        print("Footer merge A10:D10 NOT FOUND. Check logic.")
        
    # Check color
    cell = ws.cell(row=10, column=2)
    fill = cell.fill.start_color.index if cell.fill else "None"
    print(f"Cell B10 fill: {fill}")
    if fill == '00434343' or fill == 'FF434343':
        print("Color check SUCCESS")
    else:
        print(f"Color check FAILURE: expected dark gray, got {fill}")

    # Check overlaps (row counts)
    # Total rows in template was ~70. After inserting 3 and deleting 46, should be ~27.
    # But sections should be separated.
    for r in range(1, 30):
        val = ws.cell(r, 1).value
        if val: print(f"Row {r}: {val}")

except Exception as e:
    import traceback
    print(f"Error: {e}")
    traceback.print_exc()
