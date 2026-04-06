import sys, io, os, glob
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

files = sorted(glob.glob(r'temp_pdfs\*.xlsx'), key=os.path.getmtime, reverse=True)
latest = files[0]
print(f'File: {latest}')

wb = load_workbook(latest)
ws = wb.active

print()
print('=== MERGED CELLS ===')
for mr in ws.merged_cells.ranges:
    c = ws.cell(row=mr.min_row, column=mr.min_col)
    fill = c.fill.fgColor.rgb if c.fill else 'none'
    fill_type = c.fill.fill_type if c.fill else 'none'
    print(f'  Rows {mr.min_row}-{mr.max_row}, Cols {mr.min_col}-{mr.max_col}: master_fill={fill} fill_type={fill_type}')

print()
print('=== ROW 12 FULL DETAILS ===')
for col in range(1, 6):
    c = ws.cell(row=12, column=col)
    fill = c.fill.fgColor.rgb if c.fill else 'none'
    fill_type = c.fill.fill_type if c.fill else '-'
    print(f'  Col {col}: fill={fill} fill_type={fill_type}  type={type(c).__name__}')

print()
print('=== ROW 17 FULL DETAILS ===')
for col in range(1, 6):
    c = ws.cell(row=17, column=col)
    fill = c.fill.fgColor.rgb if c.fill else 'none'
    fill_type = c.fill.fill_type if c.fill else '-'
    print(f'  Col {col}: fill={fill} fill_type={fill_type} type={type(c).__name__}')
