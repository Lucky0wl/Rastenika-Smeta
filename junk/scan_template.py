import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook
ws = load_workbook('Шаблон.xlsx').active

print('=== KEY ROWS SCAN ===')
for r in range(7, 100):
    val = str(ws.cell(row=r, column=1).value or '').strip()
    if val:
        val_lower = val.lower()
        tag = ''
        if 'итого' in val_lower and 'оплате' in val_lower: tag = '<<< GRAND TOTAL'
        elif 'итого' in val_lower: tag = '<<< TOTAL'
        elif 'посадк' in val_lower or 'работ' in val_lower: tag = '<<< PLANTING'
        elif 'материал' in val_lower: tag = '<<< MATERIALS'
        elif 'налог' in val_lower or 'ндс' in val_lower: tag = '<<< TAX'
        print(f'Row {r}: {val[:60]} {tag}')
