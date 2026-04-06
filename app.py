import os
import uuid
import copy
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.cell.cell import MergedCell
from pdf_generator import PDFGenerator

app = Flask(__name__)

# Используем абсолютные пути для надежности
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
TEMP_FOLDER = os.path.join(BASE_DIR, 'temp_pdfs')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(TEMP_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['TEMP_FOLDER'] = TEMP_FOLDER

# Глобальное хранилище для загруженных данных
current_data = None
LAST_FILE_PATH = os.path.join(BASE_DIR, '.last_pricelist')

def parse_plants_from_file(filepath):
    """Парсит растения из xlsx файла. Возвращает список или None при ошибке."""
    try:
        df = pd.read_excel(filepath)
        col_map = {
            'Наименование': ['Наименование', 'Растение', 'Название', 'Name', 'Прайс', 'Товар'],
            'Кондиция': ['Кондиция', 'Параметры', 'Размер', 'Характеристики', 'Описание', 'Condition', 'Сорт'],
            'Цена': ['Цена', 'Стоимость', 'Price', 'Цена базового материала', 'Опт']
        }
        found_cols = {}
        for target, aliases in col_map.items():
            for col in df.columns:
                if any(alias.lower() in str(col).lower() for alias in aliases):
                    found_cols[target] = col
                    break
        if 'Наименование' not in found_cols:
            return None
        plants = []
        for _, row in df.iterrows():
            name = str(row[found_cols['Наименование']])
            if name == 'nan' or not name.strip(): continue
            params = str(row.get(found_cols.get('Кондиция'), '')) if 'Кондиция' in found_cols else ''
            price = row.get(found_cols.get('Цена'), 0)
            try: price = float(price) if not pd.isna(price) else 0
            except: price = 0
            plants.append({'name': name, 'parameters': params if params != 'nan' else '', 'price': price})
        return plants
    except Exception:
        return None

def _auto_load_plants():
    """Пытается загрузить последний прайс автоматически при старте сервера."""
    global current_data
    # 1. Попробовать последний использованный файл
    if os.path.exists(LAST_FILE_PATH):
        with open(LAST_FILE_PATH, 'r', encoding='utf-8') as f:
            last_path = f.read().strip()
        if os.path.exists(last_path):
            plants = parse_plants_from_file(last_path)
            if plants:
                current_data = plants
                print(f"[AutoLoad] Loaded {len(plants)} plants from {last_path}")
                return
    # 2. Если не нашли - ищем noms_38.xlsx рядом
    noms_path = os.path.join(BASE_DIR, 'noms_38.xlsx')
    if os.path.exists(noms_path):
        plants = parse_plants_from_file(noms_path)
        if plants:
            current_data = plants
            print(f"[AutoLoad] Loaded {len(plants)} plants from noms_38.xlsx")

# Выполняем автозагрузку при старте
_auto_load_plants()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    global current_data
    if 'file' not in request.files:
        return jsonify({'error': 'Файл не найден'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Файл не выбран'}), 400
    if file:
        filename = secure_filename(file.filename)
        if not filename:
            return jsonify({'error': 'Недопустимое имя файла'}), 400
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        try:
            plants = parse_plants_from_file(filepath)
            if plants is None:
                return jsonify({'error': 'Не удалось найти колонку с наименованием растений'}), 400
            current_data = plants
            # Сохраняем путь для следующего старта
            with open(LAST_FILE_PATH, 'w', encoding='utf-8') as f:
                f.write(filepath)
            return jsonify({'count': len(plants), 'items': plants})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

@app.route('/get-plants')
def get_plants():
    return jsonify(current_data or [])


def _clone_cell_style(src, dst):
    """Clone ALL style properties from src cell to dst cell using explicit openpyxl objects."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
    from openpyxl.cell.cell import MergedCell as _MergedCell
    import copy

    # MergedCell — read-only proxy, нельзя задавать стили
    if isinstance(dst, _MergedCell):
        return

    # Font
    if src.font:
        f = src.font
        color = copy.copy(f.color) if f.color else None
        dst.font = Font(
            name=f.name, size=f.size, bold=f.bold, italic=f.italic,
            underline=f.underline, strike=f.strike, color=color
        )
    
    # Fill
    if src.fill and src.fill.fill_type and src.fill.fill_type != 'none':
        fill = src.fill
        try:
            fgColor = copy.copy(fill.fgColor) if fill.fgColor else None
            bgColor = copy.copy(fill.bgColor) if fill.bgColor else None
            dst.fill = PatternFill(
                fill_type=fill.fill_type,
                fgColor=fgColor,
                bgColor=bgColor
            )
        except: pass
    
    # Alignment
    if src.alignment:
        a = src.alignment
        dst.alignment = Alignment(
            horizontal=a.horizontal, vertical=a.vertical,
            wrap_text=a.wrap_text, shrink_to_fit=a.shrink_to_fit,
            indent=a.indent, text_rotation=a.text_rotation
        )
    
    # Border - copy each side
    if src.border:
        b = src.border
        def clone_side(s):
            return Side(border_style=s.border_style, color=copy.copy(s.color)) if s else Side()
        dst.border = Border(
            left=clone_side(b.left), right=clone_side(b.right),
            top=clone_side(b.top), bottom=clone_side(b.bottom)
        )
    
    # Number format
    if src.number_format:
        dst.number_format = src.number_format


def create_app_xlsx(data):
    from openpyxl.styles import PatternFill, Alignment, Font
    template_path = os.path.join(BASE_DIR, '\u0428\u0430\u0431\u043b\u043e\u043d.xlsx')
    if not os.path.exists(template_path):
        raise FileNotFoundError('\u0424\u0430\u0439\u043b \u0448\u0430\u0431\u043b\u043e\u043d\u0430 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d')
    
    items = data.get('items', [])
    
    wb_ref = load_workbook(template_path, data_only=True)
    ws_ref = wb_ref.active
    wb = load_workbook(template_path)
    ws = wb.active
    
    # ================================================================
    # ШАГ 1: ПРЕ-СКАН шаблона (ДО любых изменений)
    # Запоминаем номера всех ключевых строк в шаблоне
    # ================================================================
    DATA_START_ROW = 7      # Первая строка данных (по шаблону)
    ref_row_height = ws_ref.row_dimensions[DATA_START_ROW].height or 18.75
    
    # \u0421\u043a\u0430\u043d\u0438\u0440\u0443\u0435\u043c \u0432\u0435\u0441\u044c \u0448\u0430\u0431\u043b\u043e\u043d \u0438 \u0437\u0430\u043f\u043e\u043c\u0438\u043d\u0430\u0435\u043c \u043f\u043e\u0437\u0438\u0446\u0438\u0438 \u0441\u0442\u0440\u043e\u043a:
    # placeholder_end = \u043f\u043e\u0441\u043b\u0435\u0434\u043d\u044f\u044f \u0441\u0442\u0440\u043e\u043a\u0430 \u0434\u0430\u043d\u043d\u044b\u0445 \u043f\u043b\u0435\u0439\u0441\u0445\u043e\u043b\u0434\u0435\u0440\u0430
    # key_rows = \u0441\u043b\u043e\u0432\u0430\u0440\u044c {'\u0438\u0442\u043e\u0433\u043e_\u0440\u0430\u0441\u0442\u0435\u043d\u0438\u0439': N, '\u0438\u0442\u043e\u0433\u043e_\u043f\u043e\u0441\u0430\u0434\u043a\u0430': M, ...}
    placeholder_end = DATA_START_ROW  # \u043f\u043e\u0441\u043b\u0435\u0434\u043d\u044f\u044f \u0441\u0442\u0440\u043e\u043a\u0430 \u043f\u043b\u0435\u0439\u0441\u0445\u043e\u043b\u0434\u0435\u0440\u0430 (\u043d\u0435 \u0432\u043a\u043b\u044e\u0447\u0430\u044f \u0438\u0442\u043e\u0433\u043e)
    
    # \u041a\u043b\u044e\u0447\u0438 \u0434\u043b\u044f \u043f\u043e\u0438\u0441\u043a\u0430 \u0441\u0442\u0440\u043e\u043a
    KEY_PATTERNS = {
        'plant_total': ['\u0438\u0442\u043e\u0433\u043e', '\u043e\u043f\u043b\u0430\u0442\u0435'],  # \u0418\u0422\u041e\u0413\u041e \u0437\u0430 \u0440\u0430\u0441\u0442\u0435\u043d\u0438\u044f
        'planting_section': ['\u043f\u043e\u0441\u0430\u0434\u043a\u0430', '\u0440\u0430\u0431\u043e\u0442'],
        'planting_total': None,  # \u0434\u043e\u043f\u043e\u043b\u043d\u044f\u0435\u043c \u043f\u043e\u0441\u043b\u0435
        'materials_section': ['\u043c\u0430\u0442\u0435\u0440\u0438\u0430\u043b'],
        'materials_total': None,  # \u0434\u043e\u043f\u043e\u043b\u043d\u044f\u0435\u043c \u043f\u043e\u0441\u043b\u0435
        'tax': ['\u043d\u0430\u043b\u043e\u0433', '\u043d\u0434\u0441'],
        'grand_total': ['\u0438\u0442\u043e\u0433\u043e \u043a \u043e\u043f\u043b\u0430\u0442\u0435', '\u0432\u0441\u0435\u0433\u043e \u043a \u043e\u043f\u043b\u0430\u0442\u0435'],
    }
    
    found_rows = {}
    in_placeholder = True
    
    for r in range(DATA_START_ROW, ws_ref.max_row + 1):
        val = str(ws_ref.cell(row=r, column=1).value or '').lower().strip()

        # Определяем конец плейсхолдеров (первая непустая строка после DATA_START_ROW с не-плейсхолдерным контентом)
        if in_placeholder:
            if not val:  # \u043f\u0443\u0441\u0442\u0430\u044f \u0441\u0442\u0440\u043e\u043a\u0430 \u0432 \u043f\u043b\u0435\u0439\u0441\u0445\u043e\u043b\u0434\u0435\u0440\u0435
                placeholder_end = r
                continue
            else:
                in_placeholder = False  # \u043d\u0430\u0448\u043b\u0438 \u0441\u0442\u0440\u043e\u043a\u0443 \u0441 \u0442\u0435\u043a\u0441\u0442\u043e\u043c - \u043a\u043e\u043d\u0435\u0446 \u043f\u043b\u0435\u0439\u0441\u0445\u043e\u043b\u0434\u0435\u0440\u043e\u0432
        
        # \u0418\u0449\u0435\u043c \u043a\u043b\u044e\u0447\u0435\u0432\u044b\u0435 \u0441\u0442\u0440\u043e\u043a\u0438
        if not val: continue
        
        if 'plant_total' not in found_rows:
            if any(k in val for k in ['\u0438\u0442\u043e\u0433\u043e', '\u043e\u043f\u043b\u0430\u0442\u0435']):
                found_rows['plant_total'] = r
        elif 'planting_section' not in found_rows:
            if any(k in val for k in ['\u043f\u043e\u0441\u0430\u0434\u043a\u0430', '\u0440\u0430\u0431\u043e\u0442']):
                found_rows['planting_section'] = r
        
    # \u042d\u0442\u043e\u0442 \u043f\u043e\u0434\u0445\u043e\u0434 \u0434\u0430\u0435\u0442 \u043d\u0430\u043c \u043f\u043e\u043b\u043d\u0443\u044e \u043a\u0430\u0440\u0442\u0443 \u0448\u0430\u0431\u043b\u043e\u043d\u0430
    # \u0422\u0435\u043f\u0435\u0440\u044c \u043f\u0440\u043e\u0441\u0442\u043e \u0441\u043a\u0430\u043d\u0438\u0440\u0443\u0435\u043c \u0432\u0435\u0441\u044c \u0448\u0430\u0431\u043b\u043e\u043d \u0438 \u0441\u043e\u0445\u0440\u0430\u043d\u044f\u0435\u043c \u0432\u0441\u0435 \u0441\u0442\u0440\u043e\u043a\u0438 \u0441 \u043a\u043b\u044e\u0447\u0435\u0432\u044b\u043c\u0438 \u0441\u043b\u043e\u0432\u0430\u043c\u0438
    key_rows = {}
    # Используем нередактированный wb (не data_only) для правильного чтения кирилицы
    # ws уже содержит оригинал шаблона (ещё не изменённый)
    # Сканируем последовательно: структура шаблона фиксирована
    scan_state = 'looking_plant_total'
    
    for r in range(DATA_START_ROW, ws.max_row + 1):
        val = str(ws.cell(row=r, column=1).value or '').lower().strip()
        if not val:
            continue

        if scan_state == 'looking_plant_total':
            # Первое ИТОГО после DATA_START_ROW = итог за растения
            if '\u0438\u0442\u043e\u0433\u043e' in val:
                key_rows['plant_total'] = r
                scan_state = 'looking_planting_section'
        
        elif scan_state == 'looking_planting_section':
            # Следующий заголовок секции = Посадка
            if '\u043f\u043e\u0441\u0430\u0434\u043a' in val or '\u0440\u0430\u0431\u043e\u0442' in val:
                scan_state = 'looking_planting_total'
        
        elif scan_state == 'looking_planting_total':
            # До ИТОГО ищем строку описания Посадки
            if 'посадка стандартная' in val or '=e10*' in str(ws.cell(row=r, column=5).value or '').lower():
                key_rows['planting_desc_row'] = r
            
            # Следующее ИТОГО = итог за посадку
            if '\u0438\u0442\u043e\u0433\u043e' in val:
                key_rows['planting_total'] = r
                scan_state = 'looking_materials_section'
        
        elif scan_state == 'looking_materials_section':
            if '\u043c\u0430\u0442\u0435\u0440\u0438\u0430\u043b' in val:
                scan_state = 'looking_materials_total'
        
        elif scan_state == 'looking_materials_total':
            if 'итого' in val:
                key_rows['materials_total'] = r
                scan_state = 'looking_delivery_section'
        
        elif scan_state == 'looking_delivery_section':
            if 'достав' in val:
                scan_state = 'looking_delivery_total'
        
        elif scan_state == 'looking_delivery_total':
            if 'итого' in val:
                if 'delivery_total' not in key_rows:
                    key_rows['delivery_total'] = r
                scan_state = 'looking_grand_total'
        
        elif scan_state == 'looking_grand_total':
            # Последнее ИТОГО = grand total (perезаписываем если несколько)
            if 'итого' in val:
                key_rows['grand_total'] = r

    
    # ================================================================
    # ШАГ 2: ПОДГОТОВКА (Очистка мержей в зоне данных)
    # Openpyxl плохо обновляет merged_cells при insert/delete_rows.
    # Мы удаляем все мержи ниже DATA_START_ROW и будем восстанавливать их вручную.
    # ================================================================
    all_merges = list(ws.merged_cells.ranges)
    for mr in all_merges:
        if mr.min_row >= DATA_START_ROW:
            try:
                ws.unmerge_cells(str(mr))
            except:
                pass

    plant_total_orig = key_rows.get('plant_total', 10)
    placeholder_count = plant_total_orig - DATA_START_ROW
    
    if placeholder_count > 0:
        ws.delete_rows(DATA_START_ROW, amount=placeholder_count)
    
    # ================================================================
    # \u0428\u0410\u0413 3: \u0412\u0441\u0442\u0430\u0432\u043b\u044f\u0435\u043c \u0441\u0442\u0440\u043e\u043a\u0438 \u0434\u0430\u043d\u043d\u044b\u0445
    # ================================================================
    n = len(items)
    if n > 0:
        ws.insert_rows(DATA_START_ROW, amount=n)
    
    # \u0421\u043c\u0435\u0449\u0435\u043d\u0438\u0435 \u0441\u0442\u0440\u043e\u043a: +n (вставили) -placeholder_count (удалили)
    offset = n - placeholder_count
    
    # Сохраняем исходные высоты ВСЕХ строк шаблона
    template_row_heights = {}
    for r in range(1, ws_ref.max_row + 1):
        h = ws_ref.row_dimensions[r].height
        if h is not None:
            template_row_heights[r] = h

    # ================================================================
    # ШАГ 4: Заполняем строки данных со стилями из шаблона
    # ================================================================
    def format_number(num):
        """Форматирует число с разделением тысяч пробелом"""
        try:
            if isinstance(num, (int, float)):
                return '{:,.2f}'.format(num).replace(',', ' ')
            return str(num)
        except:
            return str(num)

    def write_to_cell_local(row, col, val):
        from openpyxl.cell.cell import MergedCell, Cell

        target_cell = ws.cell(row=row, column=col)

        # Если это настоящая MergedCell — ищем её мастера
        if isinstance(target_cell, MergedCell):
            for mr in ws.merged_cells.ranges:
                if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                    # Форматируем числа перед вставкой
                    formatted_val = format_number(val) if col in (4, 5) else val
                    ws.cell(row=mr.min_row, column=mr.min_col).value = formatted_val
                    return

            # Если это "призрак" (MergedCell без диапазона) — уничтожаем его и создаем обычную ячейку
            if (row, col) in ws._cells:
                del ws._cells[(row, col)]
            target_cell = ws.cell(row=row, column=col)

        # Форматируем числа перед вставкой (колонки 4 и 5 - это цены и суммы)
        formatted_val = format_number(val) if col in (4, 5) else val
        # Теперь это точно обычная ячейка или мастер
        target_cell.value = formatted_val
    
    def get_ref_master(ref_row, col):
        for mr in wb_ref.active.merged_cells.ranges:
            if mr.min_row <= ref_row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                return wb_ref.active.cell(row=mr.min_row, column=mr.min_col)
        return ws_ref.cell(row=ref_row, column=col)
    
    for i, item in enumerate(items):
        row_idx = DATA_START_ROW + i
        # Авто-высота строки: Excel сам подберёт под содержимое при открытии файла.
        # Не ставим фиксированную высоту — иначе длинные названия обрезаются.
        ws.row_dimensions[row_idx].height = None

        # Сначала клонируем стили (включая wrap_text из шаблона)
        for col in range(1, 21):
            src_cell = get_ref_master(DATA_START_ROW, col)
            dst_cell = ws.cell(row=row_idx, column=col)
            if not isinstance(dst_cell, MergedCell):
                _clone_cell_style(src_cell, dst_cell)
                if col in (4, 5):
                    dst_cell.number_format = '# ##0.00'
                # Гарантируем перенос текста в колонке названий
                if col == 1:
                    from openpyxl.styles import Alignment as _Align
                    dst_cell.alignment = _Align(
                        horizontal=dst_cell.alignment.horizontal,
                        vertical=dst_cell.alignment.vertical or 'center',
                        wrap_text=True
                    )

        # Затем пишем значения (после стилей, чтобы не перезаписать number_format)
        write_to_cell_local(row_idx, 1, item.get('name'))
        write_to_cell_local(row_idx, 2, item.get('parameters', ''))
        write_to_cell_local(row_idx, 3, int(item.get('quantity', 1) or 1))
        write_to_cell_local(row_idx, 4, item.get('price', 0))
        write_to_cell_local(row_idx, 5, item.get('total'))
    
    # ================================================================
    # ШАГ 5: Заполняем футер
    # ================================================================
    def fix_footer_row(actual_row, value, style_ref_row, sum_col=5):
        from openpyxl.styles import PatternFill
        ref_height = ws_ref.row_dimensions[style_ref_row].height or 15.75
        
        def get_src_master(col):
            for mr in wb_ref.active.merged_cells.ranges:
                if mr.min_row <= style_ref_row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                    return wb_ref.active.cell(row=mr.min_row, column=mr.min_col)
            return ws_ref.cell(row=style_ref_row, column=col)
        
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row <= actual_row <= mr.max_row:
                try:
                    ws.unmerge_cells(start_row=mr.min_row, end_row=mr.max_row,
                                     start_column=mr.min_col, end_column=mr.max_col)
                except Exception:
                    pass
        
        for col in range(1, 21):
            src = get_src_master(col)
            dst = ws.cell(row=actual_row, column=col)
            _clone_cell_style(src, dst)
        
        src_master_cell = get_src_master(1)
        try:
            master_rgb = src_master_cell.fill.fgColor.rgb if src_master_cell.fill and src_master_cell.fill.fgColor else '00000000'
        except Exception:
            master_rgb = '00000000'
        if master_rgb not in ('00000000', '00FFFFFF'):
            dark_fill = PatternFill(fill_type='solid', fgColor=master_rgb)
            for col in range(1, 21):
                src = get_src_master(col)
                dst = ws.cell(row=actual_row, column=col)
                try:
                    src_fill_rgb = src.fill.fgColor.rgb if src.fill and src.fill.fgColor else '00000000'
                except Exception:
                    src_fill_rgb = '00000000'
                if src_fill_rgb not in ('00000000', '00FFFFFF', None):
                    dst.fill = PatternFill(fill_type='solid', fgColor=src_fill_rgb)
        
        ws.row_dimensions[actual_row].height = ref_height
        
        if value is not None:
            write_to_cell_local(actual_row, sum_col, value)
            ws.cell(row=actual_row, column=sum_col).number_format = '# ##0.00'
        
        try:
            ws.merge_cells(start_row=actual_row, end_row=actual_row, start_column=1, end_column=sum_col-1)
            align = Alignment(horizontal='right', vertical='center')
            ws.cell(row=actual_row, column=1).alignment = align
        except:
            pass
    
    calc_plant_total = sum(float(i.get('price') or 0) * float(i.get('quantity') or 0) for i in items)
    extra_mat_total = float(data.get('material_total', 0) or 0)
    tax_rate = float(data.get('tax_rate', 0) or 0)
    tax_amount = calc_plant_total * (tax_rate / 100)
    
    FOOTER_SECTION_ROW = plant_total_orig
    FOOTER_GRAND_ROW = 25

    materials = data.get('materials', [])
    m = len(materials)

    # Сохраняем сдвиг ДО обработки материалов.
    # plant_offset применяется к строкам ВЫШЕ секции Материалов
    # (Растения ИТОГО, Посадка описание, Посадка ИТОГО).
    # После обработки материалов offset дополнительно изменится на (m - dummy_count),
    # и этот полный offset применяется к строкам AT/BELOW секции Материалов.
    plant_offset = offset

    if 'materials_total' in key_rows:
        mat_total_row = key_rows['materials_total']
        mat_header_row = mat_total_row - 1
        while mat_header_row > 1 and str(ws_ref.cell(row=mat_header_row, column=1).value or '').strip() != 'Наименование':
            mat_header_row -= 1
        dummy_count = mat_total_row - mat_header_row - 1
        actual_insert_start = mat_header_row + plant_offset + 1
        ws.delete_rows(actual_insert_start, amount=dummy_count)
        if m > 0:
            ws.insert_rows(actual_insert_start, amount=m)
            for i, mat in enumerate(materials):
                row_idx = actual_insert_start + i
                ws.row_dimensions[row_idx].height = None
                for col in range(1, 21):
                    src_ref = get_ref_master(key_rows['materials_total'] - 1, col)
                    dst = ws.cell(row=row_idx, column=col)
                    if not isinstance(dst, MergedCell):
                        _clone_cell_style(src_ref, dst)
                        if col in (4, 5):
                            dst.number_format = '# ##0.00'
                        if col == 1:
                            from openpyxl.styles import Alignment as _Align
                            dst.alignment = _Align(
                                horizontal=dst.alignment.horizontal,
                                vertical=dst.alignment.vertical or 'center',
                                wrap_text=True
                            )
                write_to_cell_local(row_idx, 1, mat.get('name'))
                write_to_cell_local(row_idx, 2, mat.get('parameters', ''))
                write_to_cell_local(row_idx, 3, int(mat.get('quantity', 1) or 1))
                write_to_cell_local(row_idx, 4, mat.get('price', 0))
                write_to_cell_local(row_idx, 5, mat.get('total'))
        offset = plant_offset + (m - dummy_count)

    # --- Строки ВЫШЕ секции Материалов — используем plant_offset ---
    if 'plant_total' in key_rows:
        fix_footer_row(key_rows['plant_total'] + plant_offset, calc_plant_total, style_ref_row=FOOTER_SECTION_ROW)

    labor_total = float(data.get('labor_total', 0) or 0)
    if 'planting_total' in key_rows:
        fix_footer_row(key_rows['planting_total'] + plant_offset, labor_total, style_ref_row=FOOTER_SECTION_ROW)
        if 'planting_desc_row' in key_rows:
            p_desc_row = key_rows['planting_desc_row'] + plant_offset
            ws.row_dimensions[p_desc_row].height = 50
            write_to_cell_local(p_desc_row, 5, labor_total)
            target_cell = ws.cell(row=p_desc_row, column=5)
            if not isinstance(target_cell, MergedCell):
                target_cell.number_format = '# ##0.00'

    # --- Строки AT/BELOW секции Материалов — используем полный offset ---
    if 'materials_total' in key_rows:
        fix_footer_row(key_rows['materials_total'] + offset, extra_mat_total, style_ref_row=FOOTER_SECTION_ROW)

    extra_delivery_total = float(data.get('delivery_total', 0) or 0)
    if 'delivery_total' in key_rows:
        delivery_data_row = key_rows['delivery_total'] + offset - 1
        if delivery_data_row > 0:
            for col in range(1, 6):
                cell = ws.cell(row=delivery_data_row, column=col)
                if not isinstance(cell, MergedCell):
                    if cell.value and str(cell.value).startswith('='):
                        cell.value = None
                else:
                    for mr in ws.merged_cells.ranges:
                        if mr.min_row <= delivery_data_row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                            m_cell = ws.cell(row=mr.min_row, column=mr.min_col)
                            if m_cell.value and str(m_cell.value).startswith('='):
                                m_cell.value = None

            write_to_cell_local(delivery_data_row, 5, extra_delivery_total)
            target_cell = ws.cell(row=delivery_data_row, column=5)
            if not isinstance(target_cell, MergedCell):
                target_cell.number_format = '# ##0.00'

    # Убираем лишние промежуточные итоговые строки по просьбе пользователя.
    # Чтобы не сбились индексы, удаляем снизу вверх или корректируем финальный индекс.
    rows_to_delete = []
    if 'delivery_total' in key_rows: rows_to_delete.append(key_rows['delivery_total'] + offset)
    if 'tax' in key_rows: rows_to_delete.append(key_rows['tax'] + offset)
    rows_to_delete.sort(reverse=True)

    final_grand_total_row = key_rows.get('grand_total', 26) + offset
    for r in rows_to_delete:
        ws.delete_rows(r)
        if r < final_grand_total_row:
            final_grand_total_row -= 1

    if 'grand_total' in key_rows:
        grand_total = float(data.get('grand_total', 0) or 0)
        fix_footer_row(final_grand_total_row, grand_total, style_ref_row=FOOTER_GRAND_ROW)

    # ================================================================
    # ФИНАЛЬНЫЙ ПРОХОД: для каждого merged range:
    #   3) re-merge (восстанавливает центрирование текста)
    # Excel при отрисовке merged range берёт fill мастера — он правильный.
    # Slave cells становятся proxy после re-merge, но это не важно для Excel.
    # ================================================================
    from openpyxl.styles import PatternFill, Alignment
    
    all_merges = list(ws.merged_cells.ranges)
    
    for mr in all_merges:
        master_row, master_col = mr.min_row, mr.min_col
        # ПРОПУСКАЕМ ТОЛЬКО ПЕРВУЮ СТРОКУ (лого), остальные (Комм. предложение, контакты) 
        # должны быть в проходе для заливки slave ячеек, иначе они будут белыми.
        if master_row < 2:
            continue
            
        r1, r2, c1, c2 = mr.min_row, mr.max_row, mr.min_col, mr.max_col
        master = ws.cell(row=master_row, column=master_col)
        
        # Сохраняем стиль мастера
        master_fill = master.fill
        master_align = master.alignment
        master_font = master.font
        
        if master_fill and master_fill.fill_type and master_fill.fill_type not in (None, 'none'):
            try:
                solid_fill = PatternFill(fill_type='solid', fgColor=master_fill.fgColor.rgb)
            except Exception:
                solid_fill = None
        else:
            solid_fill = None
        
        # Шаг 1: unmerge
        try:
            ws.unmerge_cells(start_row=r1, end_row=r2, start_column=c1, end_column=c2)
        except Exception:
            continue
        
        # Шаг 2: применяем fill ко всем ячейкам
        if solid_fill:
            for row in range(r1, r2 + 1):
                for col in range(c1, c2 + 1):
                    c = ws.cell(row=row, column=col)
                    c.fill = solid_fill
        
        # Явно восстанавливаем alignment и font на мастере перед re-merge
        if master_align:
            master.alignment = copy.copy(master_align)
        if master_font:
            master.font = copy.copy(master_font)
        
        # Шаг 3: re-merge — восстанавливает центрирование текста
        try:
            ws.merge_cells(start_row=r1, end_row=r2, start_column=c1, end_column=c2)
        except Exception:
            pass

    # ================================================================
    # ДОПОЛНИТЕЛЬНЫЙ ПРОХОД: финальная зачистка белых полей
    # Проходим по всем строкам: если Col A имеет заливку, красим всю строку B-E
    # ================================================================
    for row_idx in range(1, ws.max_row + 1):
        c_a = ws.cell(row=row_idx, column=1)
        if c_a.fill and c_a.fill.fill_type == 'solid':
            a_rgb = str(c_a.fill.fgColor.rgb or '00000000').upper()
            if a_rgb in ('00000000', 'FFFFFFFF', '00FFFFFF', 'NONE'):
                continue
            
            # Подготавливаем заливку
            row_fill = PatternFill(fill_type='solid', fgColor=a_rgb)
            
            # Если строка была объединена (как Наименование или заголовок), 
            # openpyxl может капризничать — временно разрываем объединение для заливки
            row_merges = [mr for mr in list(ws.merged_cells.ranges) if mr.min_row <= row_idx <= mr.max_row]
            for mr in row_merges:
                try: ws.unmerge_cells(start_row=mr.min_row, end_row=mr.max_row, start_column=mr.min_col, end_column=mr.max_col)
                except: pass
            
            # Закрашиваем реальные ячейки
            for col_idx in range(1, 6):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill
            
            # Возвращаем объединение
            for mr in row_merges:
                try: ws.merge_cells(start_row=mr.min_row, end_row=mr.max_row, start_column=mr.min_col, end_column=mr.max_col)
                except: pass

    # ================================================================
    # ВОССТАНАВЛИВАЕМ ВЫСОТЫ СТРОК (openpyxl теряет их при insert_rows)
    # ================================================================
    for r_orig, h in template_row_heights.items():
        if r_orig < DATA_START_ROW:
            ws.row_dimensions[r_orig].height = h

    # Жестко проставим высоты и мержи для ключевых строк шаблона с большим шрифтом:
    for r in range(1, ws.max_row + 1):
        cell_val = str(ws.cell(row=r, column=1).value or '').strip()
        
        # 1. Заголовки разделов (Растения, Посадка и т.д.)
        if cell_val in ['Растения', 'Посадка', 'Материалы', 'Доставка']:
            try: ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=5)
            except: pass
            ws.row_dimensions[r].height = 36 
            c = ws.cell(row=r, column=1)
            from openpyxl.styles import Font, Alignment
            c.font = Font(name='Manrope', size=20, bold=True, color='FFFFFFFF')
            c.alignment = Alignment(horizontal='left', vertical='center')

        # 2. Подзаголовки (Наименование) — часто в шаблоне они на всю ширину (или до Суммы)
        elif cell_val.lower() == 'наименование':
            # Если это секция Растения — Col 5 это Сумма. Если Посадка — тоже.
            # Попробуем мержить A:D если Col 5 имеет заголовок 'Сумма'
            col5_val = str(ws.cell(row=r, column=5).value or '').lower()
            if 'сумма' in col5_val:
                # В Растениях Col 2,3,4 имеют свои заголовки, не мержим
                pass
            else:
                # В Посадке Наименование обычно на почти всю ширину
                try: ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=4)
                except: pass

        # 3. Строка с описанием посадки (начинается с "Посадка стандартная")
        elif cell_val.startswith('Посадка стандартная'):
            try: ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=4)
            except: pass
            ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

    filename = f"estimate_{uuid.uuid4()}.xlsx"
    filepath = os.path.join(app.config['TEMP_FOLDER'], filename)
    wb.save(filepath)
    return filepath


@app.route('/generate-xlsx', methods=['POST'])
def generate_xlsx():
    data = request.json
    if not data:
        return jsonify({'error': 'Данные не получены'}), 400
    try:
        filepath = create_app_xlsx(data)
        return send_file(os.path.abspath(filepath), as_attachment=True, download_name="Коммерческое_предложение_Ландшафт.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/generate-pdf', methods=['POST'])
def generate_pdf():
    data = request.json
    if not data: return jsonify({'error': 'Данные не получены'}), 400
    try:
        from datetime import datetime
        from pdf_generator import PDFGenerator
        
        # 1. Подготовка данных для шаблона
        template_data = {
            'date': datetime.now().strftime("%d.%m.%Y"),
            'items': data.get('items', []),
            'materials': data.get('materials', []),
            'sum_materials': float(data.get('sum_materials', 0) or 0),
            'labor_total': float(data.get('labor_total', 0) or 0),
            'delivery_total': float(data.get('delivery_total', 0) or 0),
            'tax_rate': float(data.get('tax_rate', 0) or 0),
            'sum_tax': float(data.get('sum_tax', 0) or 0),
            'grand_total': float(data.get('grand_total', 0) or 0),
            'company_name': data.get('company_name', 'Rastenika'),
            'company_contacts': data.get('company_contacts', '')
        }
        
        # 2. Рендерим HTML
        html_content = render_template('estimate_pdf.html', **template_data)
        
        # 3. Генерируем PDF через Playwright
        pdf_gen = PDFGenerator()
        pdf_path = pdf_gen.create_pdf_from_html(html_content)
        
        return send_file(os.path.abspath(pdf_path), as_attachment=True, download_name="Коммерческое_предложение_Ландшафт.pdf", mimetype='application/pdf')
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
